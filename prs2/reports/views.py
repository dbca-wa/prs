from django.conf import settings
from django.urls import reverse
from django.http import HttpResponse
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Font
from taggit.models import Tag

from referral.models import Referral, Clearance, Task, TaskType
from referral.utils import breadcrumbs_li, is_model_or_string, prs_user


class ReportView(TemplateView):
    """Template view to allow filtering and download of data.
    """
    template_name = 'reports/reports.html'

    def get_context_data(self, **kwargs):
        context = super(ReportView, self).get_context_data(**kwargs)
        context['page_title'] = ' | '.join([settings.APPLICATION_ACRONYM, 'Reports'])
        links = [(reverse('site_home'), 'Home'), (None, 'Reports')]
        context['breadcrumb_trail'] = breadcrumbs_li(links)
        context['no_sidebar'] = True
        context['is_prs_user'] = prs_user(self.request)
        return context


class DownloadView(TemplateView):
    """A basic view to return a spreadsheet of referral objects.
    """
    def dispatch(self, request, *args, **kwargs):
        # kwargs must include a Model class, or a string.
        # Determine the required model type.
        if 'model' in kwargs:
            self.model = is_model_or_string(kwargs.pop('model'))
        return super(DownloadView, self).dispatch(request, *args, **kwargs)

    def get(self, request):
        # Get any query parameters to filter the data.
        query_params = dict(request.GET.items())
        # Get the required model type from the query params.
        model = is_model_or_string(query_params.pop('model'))
        # Special case: region -> regions.
        region = query_params.pop('region__id', None)
        if region:
            query_params['regions__id__in'] = [region]
        # Special case: remove tag PKs from the query params.
        tag = query_params.pop('tag__id', None)
        if tag:
            tags = Tag.objects.filter(pk=tag)
        else:
            tags = None

        # Generate a blank Excel workbook.
        wb = Workbook()
        ws = wb.active  # The worksheet
        # Default font for all cells.
        arial = Font(name='Arial', size=10)
        # Define a date style.
        date_style = 'dd/mm/yyyy'

        # Generate a HTTPResponse object to write to.
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        if model == Referral:
            response['Content-Disposition'] = 'attachment; filename=prs_referrals.xlsx'
            # Filter referral objects according to the parameters.
            referrals = Referral.objects.current().filter(**query_params)
            if tags:  # Optional: filter by tags.
                referrals = referrals.filter(tags__in=tags).distinct()
            # Write the column headers to the new worksheet.
            headers = [
                'Referral ID', 'Region(s)', 'Referrer', 'Type', 'Reference',
                'Received', 'Description', 'Address', 'Triggers', 'Tags',
                'File no.', 'LGA']
            for col, value in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = value
                cell.font = arial
            # Write the referral values to the worksheet.
            for row, r in enumerate(referrals, 2):  # Start at row 2
                cell = ws.cell(row=row, column=1)
                cell.value = r.pk
                cell.font = arial
                cell = ws.cell(row=row, column=2)
                cell.value = r.regions_str
                cell.font = arial
                cell = ws.cell(row=row, column=3)
                cell.value = r.referring_org.name
                cell.font = arial
                cell = ws.cell(row=row, column=4)
                cell.value = r.type.name
                cell.font = arial
                cell = ws.cell(row=row, column=5)
                cell.value = r.reference
                cell.font = arial
                cell = ws.cell(row=row, column=6)
                cell.value = r.referral_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=7)
                cell.value = r.description
                cell.font = arial
                cell = ws.cell(row=row, column=8)
                cell.value = r.address
                cell.font = arial
                cell = ws.cell(row=row, column=9)
                cell.value = ', '.join([t.name for t in r.dop_triggers.all()])
                cell.font = arial
                cell = ws.cell(row=row, column=10)
                cell.value = ', '.join([t.name for t in r.tags.all()])
                cell.font = arial
                cell = ws.cell(row=row, column=11)
                cell.value = r.file_no
                cell.font = arial
                cell = ws.cell(row=row, column=12)
                cell.value = r.lga.name if r.lga else ''
                cell.font = arial
        elif model == Clearance:
            response['Content-Disposition'] = 'attachment; filename=prs_clearance_requests.xlsx'
            # Filter clearance objects according to the parameters.
            clearances = Clearance.objects.current().filter(**query_params)
            # Write the column headers to the new worksheet.
            headers = [
                'Referral ID', 'Region(s)', 'Reference', 'Condition no.',
                'Approved condition', 'Category', 'Task description',
                'Deposited plan no.', 'Assigned user', 'Status', 'Start date',
                'Due date', 'Complete date', 'Stop date', 'Restart date',
                'Total stop days']
            for col, value in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = value
                cell.font = arial
            # Write the clearance values to the worksheet.
            for row, c in enumerate(clearances, 2):  # Start at row 2
                cell = ws.cell(row=row, column=1)
                cell.value = c.condition.referral.pk
                cell.font = arial
                cell = ws.cell(row=row, column=2)
                cell.value = c.condition.referral.regions_str
                cell.font = arial
                cell = ws.cell(row=row, column=3)
                cell.value = c.condition.referral.reference
                cell.font = arial
                cell = ws.cell(row=row, column=4)
                cell.value = c.condition.identifier
                cell.font = arial
                cell = ws.cell(row=row, column=5)
                cell.value = c.condition.condition
                cell.font = arial
                cell = ws.cell(row=row, column=6)
                cell.font = arial
                if c.condition.category:
                    cell.value = c.condition.category.name
                cell = ws.cell(row=row, column=7)
                cell.value = c.task.description
                cell.font = arial
                cell = ws.cell(row=row, column=8)
                cell.value = c.deposited_plan
                cell.font = arial
                cell = ws.cell(row=row, column=9)
                cell.value = c.task.assigned_user.get_full_name()
                cell.font = arial
                cell = ws.cell(row=row, column=10)
                cell.value = c.task.state.name
                cell.font = arial
                cell = ws.cell(row=row, column=11)
                cell.value = c.task.start_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=12)
                cell.value = c.task.due_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=13)
                cell.value = c.task.complete_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=14)
                cell.value = c.task.stop_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=15)
                cell.value = c.task.restart_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=16)
                cell.value = c.task.stop_time
                cell.font = arial
        elif model == Task:
            response['Content-Disposition'] = 'attachment; filename=prs_tasks.xlsx'
            # Filter task objects according to the parameters.
            tasks = Task.objects.current().filter(**query_params)
            # Business rule: filter out 'Condition clearance' task types.
            cr = TaskType.objects.get(name='Conditions clearance request')
            tasks = tasks.exclude(type=cr)
            # Write the column headers to the new worksheet.
            headers = [
                'Task ID', 'Region(s)', 'Referral ID', 'Referred by',
                'Referral type', 'Reference', 'Referral received', 'Task type',
                'Task status', 'Assigned user', 'Task start', 'Task due',
                'Task complete', 'Stop date', 'Restart date', 'Total stop days',
                'File no.', 'DoP triggers', 'Referral description',
                'Referral address', 'LGA']
            for col, value in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = value
                cell.font = arial
            # Write the task values to the worksheet.
            for row, t in enumerate(tasks, 2):  # Start at row 2
                cell = ws.cell(row=row, column=1)
                cell.value = t.pk
                cell.font = arial
                cell = ws.cell(row=row, column=2)
                cell.value = t.referral.regions_str
                cell.font = arial
                cell = ws.cell(row=row, column=3)
                cell.value = t.referral.pk
                cell.font = arial
                cell = ws.cell(row=row, column=4)
                cell.value = t.referral.referring_org.name
                cell.font = arial
                cell = ws.cell(row=row, column=5)
                cell.value = t.referral.type.name
                cell.font = arial
                cell = ws.cell(row=row, column=6)
                cell.value = t.referral.reference
                cell.font = arial
                cell = ws.cell(row=row, column=7)
                cell.value = t.referral.referral_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=8)
                cell.value = t.type.name
                cell.font = arial
                cell = ws.cell(row=row, column=9)
                cell.value = t.state.name
                cell.font = arial
                cell = ws.cell(row=row, column=10)
                cell.value = t.assigned_user.get_full_name()
                cell.font = arial
                cell = ws.cell(row=row, column=11)
                cell.value = t.start_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=12)
                cell.value = t.due_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=13)
                cell.value = t.complete_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=14)
                cell.value = t.stop_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=15)
                cell.value = t.restart_date
                cell.number_format = date_style
                cell.font = arial
                cell = ws.cell(row=row, column=16)
                cell.value = t.stop_time
                cell.font = arial
                cell = ws.cell(row=row, column=17)
                cell.value = t.referral.file_no
                cell.font = arial
                cell = ws.cell(row=row, column=18)
                cell.value = ', '.join([i.name for i in t.referral.dop_triggers.all()])
                cell.font = arial
                cell = ws.cell(row=row, column=19)
                cell.value = t.referral.description
                cell.font = arial
                cell = ws.cell(row=row, column=20)
                cell.value = t.referral.address
                cell.font = arial
                cell = ws.cell(row=row, column=21)
                cell.value = t.referral.lga.name if t.referral.lga else ''
                cell.font = arial

        wb.save(response)  # Save the workbook contents to the response.
        return response
